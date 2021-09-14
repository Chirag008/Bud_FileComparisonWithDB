import os
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import PatternFill


class Xlsx_File_Handler:
    xlsx_file = None
    xlsx_sheet = None
    absolute_file_path = None
    row_index = 1
    header_to_col_number_map = {}
    unmatched_data_fill_style = None

    def __init__(self, file_name, headers_list):
        self.absolute_file_path = self.get_complete_xlsx_file_path(file_name)
        self.xlsx_file = Workbook()
        self.xlsx_sheet = self.xlsx_file.active
        self.xlsx_sheet.title = 'Comparison_Result'
        self.number_of_cols = len(headers_list)
        self.write_row_in_xlsx_file(headers_list)
        for index, col_name in enumerate(headers_list, start=1):
            self.header_to_col_number_map[col_name] = index
        self.unmatched_data_fill_style = PatternFill(start_color='FFFFA500',
                                                     end_color='FFFFA500',
                                                     fill_type='solid')


    def get_complete_xlsx_file_path(self, file_name):
        project_root = Path(os.path.abspath(os.path.dirname(__file__))).parent
        file_dir = 'report/xlsx_output'
        file_dir = os.path.join(project_root, file_dir)
        if not os.path.exists(file_dir):
            os.makedirs(file_dir)
        file_path = os.path.join(file_dir, file_name)
        return file_path

    def write_row_in_xlsx_file(self, list_of_col_values_for_a_row: list):
        for col_num, val in enumerate(list_of_col_values_for_a_row, start=1):
            self.xlsx_sheet.cell(row=self.row_index, column=col_num).value = val
        self.row_index += 1

    def write_dict_as_row_in_xlsx_file(self, row_as_dictionary: dict, unmatched_col_names_as_list: list = None):
        for col, val in row_as_dictionary.items():
            self.xlsx_sheet.cell(row=self.row_index, column=self.header_to_col_number_map[col]).value = val
        if unmatched_col_names_as_list:
            for col in unmatched_col_names_as_list:
                self.xlsx_sheet.cell(row=self.row_index,
                                     column=self.header_to_col_number_map[col]).fill = self.unmatched_data_fill_style
        self.row_index += 1

    def save_xlsx_file(self):
        # check if workbook is opened by some other process. Close the workbook in that case

        print('Saving xlsx file ... ')
        self.xlsx_file.save(filename=self.absolute_file_path)
        self.xlsx_file.close()
        print('XLSX file saved successfully!!')


if __name__ == '__main__':
    fh = Xlsx_File_Handler('testfile.xlsx', ['col1', 'col2', 'col3'])
    fh.save_xlsx_file()
